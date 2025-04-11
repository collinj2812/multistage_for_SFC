import pandas as pd
import pickle
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.title import Title
from typing import Tuple, Optional, Dict
import os
import numpy as np


def save_excel(df, output_path, metadata=None):
    """
    Create an Excel file with plots and metadata from a pandas DataFrame.

    Parameters:
    df (pandas.DataFrame): Input DataFrame with time series data
    output_path (str): Path where the Excel file will be saved
    metadata (dict): Dictionary containing metadata to be added to the Excel file
    """
    # First save the DataFrame to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Data', index=True)

        # Get the workbook
        workbook = writer.book

        # Create a new sheet for plots
        plot_sheet = workbook.create_sheet('Plots')

        # Create a new sheet for metadata
        if metadata:
            meta_sheet = workbook.create_sheet('Metadata')
            row = 1
            for key, value in metadata.items():
                meta_sheet.cell(row=row, column=1, value=key)
                meta_sheet.cell(row=row, column=2, value=str(value))
                row += 1

        # Create plots for each numeric column
        numeric_cols = df.select_dtypes(include=['float64', 'int64']).columns
        for i, col in enumerate(numeric_cols):
            # Create a line chart
            chart = LineChart()
            chart.title = col
            chart.style = 2
            chart.height = 10
            chart.width = 20

            # Define data references
            data = Reference(
                workbook['Data'],
                min_col=df.columns.get_loc(col) + 2,  # +2 because Excel is 1-based and we have index
                min_row=1,
                max_row=len(df) + 1
            )

            # Define categories (x-axis values)
            categories = Reference(
                workbook['Data'],
                min_col=1,
                min_row=2,
                max_row=len(df) + 1
            )

            # Add data and categories to chart
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)

            # Position charts in a grid (2 columns)
            row_pos = (i // 2) * 15
            col_pos = (i % 2) * 15

            # Add chart to plot sheet
            plot_sheet.add_chart(chart, f"{chr(65 + col_pos)}{row_pos + 1}")

def read_excel(file_path: str) -> Tuple[pd.DataFrame, Optional[Dict]]:
    """
    Read an Excel file and return both the data and metadata if available.

    Parameters:
    file_path (str): Path to the Excel file

    Returns:
    Tuple[pd.DataFrame, Optional[Dict]]:
        - The main data as a DataFrame
        - Dictionary of metadata (if metadata sheet exists, else None)

    Raises:
    FileNotFoundError: If the specified file doesn't exist
    ValueError: If the 'Data' sheet is not found in the Excel file
    """
    try:
        # Read the main data
        data_df = pd.read_excel(
            file_path,
            sheet_name='Data',
            index_col=0  # Assuming the first column is the index
        )

        # Try to read metadata if it exists
        try:
            metadata_df = pd.read_excel(file_path, sheet_name='Metadata', header=None)
            # Convert metadata DataFrame to dictionary
            metadata = dict(zip(metadata_df[0], metadata_df[1]))
        except ValueError:
            # No metadata sheet found
            metadata = None

        return data_df, metadata

    except FileNotFoundError:
        raise FileNotFoundError(f"The file {file_path} was not found.")
    except ValueError as e:
        if "No sheet named 'Data'" in str(e):
            raise ValueError(
                "This Excel file doesn't contain a 'Data' sheet. Make sure you're reading the correct file.")
        raise e


def save_sfcmodel(model, folder_path):
    """
    Save an SFCModel object to a folder structure with simple text files.

    Args:
        model: The SFCModel object to save
        folder_path: Path to the output folder
    """
    # Create the main folder if it doesn't exist
    os.makedirs(folder_path, exist_ok=True)

    # Get model attributes
    if hasattr(model, '__dict__'):
        attributes = vars(model)
    elif isinstance(model, dict):
        attributes = model
    else:
        attributes = {attr: getattr(model, attr) for attr in dir(model)
                      if not attr.startswith('__') and not callable(getattr(model, attr))}

    # Save each top-level attribute
    for key, value in attributes.items():
        # Skip protected attributes
        if key.startswith('_'):
            continue

        if isinstance(value, dict):
            # Create a subfolder for dictionaries
            dict_folder = os.path.join(folder_path, key)
            os.makedirs(dict_folder, exist_ok=True)

            # Save each item in the dictionary
            for k, v in value.items():
                _save_value(v, os.path.join(dict_folder, str(k)))

        elif isinstance(value, list) or isinstance(value, np.ndarray):
            # Handle lists and arrays
            if len(value) > 0 and (hasattr(value[0], '__dict__') or isinstance(value[0], dict)):
                # List of objects or dictionaries
                list_folder = os.path.join(folder_path, key)
                os.makedirs(list_folder, exist_ok=True)

                # Save each item in the list
                for i, item in enumerate(value):
                    item_path = os.path.join(list_folder, str(i))
                    _save_value(item, item_path)
            else:
                # Simple list or array
                _save_value(value, os.path.join(folder_path, key))

        else:
            # Simple value or object
            _save_value(value, os.path.join(folder_path, key))

    print(f"Model successfully saved to {folder_path}")


def _save_value(value, path):
    """Helper function to save a value to a text file."""
    if isinstance(value, dict):
        # Create a folder for the dictionary
        os.makedirs(path, exist_ok=True)

        # Save each item in the dictionary
        for k, v in value.items():
            _save_value(v, os.path.join(path, str(k)))

    elif isinstance(value, list) or isinstance(value, np.ndarray):
        if len(value) > 0 and (hasattr(value[0], '__dict__') or isinstance(value[0], dict)):
            # List of objects or dictionaries
            os.makedirs(path, exist_ok=True)

            # Save each item in the list
            for i, item in enumerate(value):
                _save_value(item, os.path.join(path, str(i)))
        else:
            # For arrays, save as text
            with open(path + '.txt', 'w') as f:
                if isinstance(value, np.ndarray):
                    # Save shape information for arrays
                    f.write(f"#SHAPE: {','.join(str(dim) for dim in value.shape)}\n")

                    # Handle multi-dimensional arrays
                    if value.ndim > 1:
                        for i, row in enumerate(value):
                            f.write(f"#ROW {i}\n")
                            if isinstance(row, np.ndarray):
                                f.write(' '.join(str(item) for item in row) + '\n')
                            else:
                                f.write(str(row) + '\n')
                    else:
                        # 1D array
                        f.write(' '.join(str(item) for item in value) + '\n')
                else:
                    # Regular list
                    f.write(' '.join(str(item) for item in value) + '\n')

    elif hasattr(value, '__dict__'):
        # Create a folder for the object
        os.makedirs(path, exist_ok=True)

        # Save each attribute of the object
        for k, v in vars(value).items():
            if not k.startswith('_'):  # Skip protected attributes
                _save_value(v, os.path.join(path, k))

    else:
        # Simple value, save as text
        with open(path + '.txt', 'w') as f:
            f.write(str(value))


def load_sfcmodel(folder_path):
    """
    Load an SFCModel from a folder structure into a dictionary.

    Args:
        folder_path: Path to the input folder

    Returns:
        A dictionary containing the model data
    """
    if not os.path.exists(folder_path):
        raise FileNotFoundError(f"Folder not found: {folder_path}")

    model_dict = {}

    # Get all items in the main folder
    items = os.listdir(folder_path)

    for item in items:
        item_path = os.path.join(folder_path, item)

        # Get the key name (remove file extension if present)
        key = item.split('.')[0]

        if os.path.isdir(item_path):
            # This is a folder, could be a dict, list, or object
            model_dict[key] = _load_value(item_path)

        elif item.endswith('.txt'):
            # Simple value or array in text file
            model_dict[key] = _load_value(item_path)

    print(f"Model successfully loaded from {folder_path}")
    return model_dict


def _load_value(path):
    """Helper function to load a value from a file or folder."""
    if os.path.isdir(path):
        # This is a folder, check if it's a numbered list
        items = os.listdir(path)

        # Check if all items are numbered
        numbered_items = [item for item in items if item.split('.')[0].isdigit()]

        if len(numbered_items) > 0 and len(numbered_items) == len(items):
            # This is a list
            result_list = []

            # Sort by number
            numbered_items.sort(key=lambda x: int(x.split('.')[0]))

            for item in numbered_items:
                item_path = os.path.join(path, item)
                item_index = int(item.split('.')[0])

                # Ensure the list is long enough
                while len(result_list) <= item_index:
                    result_list.append(None)

                # Load the item
                result_list[item_index] = _load_value(item_path)

            return result_list

        else:
            # This is a dict or object
            result_dict = {}

            for item in items:
                item_path = os.path.join(path, item)
                key = item.split('.')[0]

                # Load the item
                result_dict[key] = _load_value(item_path)

            return result_dict

    elif path.endswith('.txt'):
        # Read the text file
        with open(path, 'r') as f:
            lines = f.readlines()

        if not lines:
            return None

        # Check if it's an array
        if len(lines) > 0 and lines[0].startswith('#SHAPE:'):
            # This is a numpy array
            shape_info = lines[0].strip().replace('#SHAPE:', '').strip()
            shape = tuple(int(dim) for dim in shape_info.split(',') if dim)

            if len(shape) > 1:
                # Multi-dimensional array
                data = []
                current_row = []

                for line in lines[1:]:
                    if line.startswith('#ROW'):
                        if current_row:
                            data.append(current_row)
                            current_row = []
                    else:
                        values = line.strip().split()
                        if values:
                            current_row = [_parse_value(val) for val in values]

                if current_row:
                    data.append(current_row)

                return np.array(data)
            else:
                # 1D array
                if len(lines) > 1:
                    values = lines[1].strip().split()
                    return np.array([_parse_value(val) for val in values])
                else:
                    return np.array([])

        # Check if it contains multiple values (might be a list)
        first_line = lines[0].strip()
        if ' ' in first_line and not first_line.startswith('#'):
            # This might be a list
            values = first_line.split()
            return [_parse_value(val) for val in values]

        # Simple value
        return _parse_value(lines[0].strip())

    else:
        # Check if there's a text file with this name
        txt_path = path + '.txt'
        if os.path.exists(txt_path):
            return _load_value(txt_path)

        # No file found
        return None


def _parse_value(value_str):
    """Parse a string value into the appropriate Python type."""
    # Try to parse as an integer
    try:
        return int(value_str)
    except ValueError:
        pass

    # Try to parse as a float
    try:
        return float(value_str)
    except ValueError:
        pass

    # Check for boolean values
    if value_str.lower() == 'true':
        return True
    elif value_str.lower() == 'false':
        return False
    elif value_str.lower() == 'none':
        return None

    # Default to string
    return value_str


def load_simulation_data(model_type, load_dir="results"):
    """
    Load simulation data from pickle file.

    Args:
        model_type: Type of model ('nn', 'bll', or 'cqr')
        load_dir: Base directory for loading results

    Returns:
        dict: Dictionary containing the simulation data
    """
    data_path = os.path.join(load_dir, model_type, f'{model_type}_simulation_data.pkl')
    try:
        with open(data_path, 'rb') as f:
            data = pickle.load(f)
        print(f"Data loaded from {data_path}")
        return data
    except FileNotFoundError:
        print(f"No data file found at {data_path}")
        return None